"""
enrich_hikvision_catalog.py
----------------------------
Extractor externo para enriquecer una lista de SKU de Hikvision
usando páginas públicas oficiales de hikvision.com.

Usa navegador real (Playwright + Chromium) para renderizar páginas dinámicas.
No requiere credenciales, cookies ni tokens. Descarga imágenes y metadatos
cuando el SKU se confirma de forma estricta contra contenido oficial.

Autor: Grupo Security — Herramienta de línea de comandos
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
import re
import sys
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional
from urllib.parse import urljoin, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from bs4.element import Tag

try:
    from playwright.sync_api import sync_playwright
except ImportError:
    sync_playwright = None  # type: ignore


# --- Constantes ---

ALLOWED_DOMAINS = frozenset({
    "www.hikvision.com",
    "hikvision.com",
    "display.hikvision.com",
    "pro-av.hikvision.com",
})

USER_AGENT = "GrupoSecurityCatalogEnricher/1.0 (authorized-catalog-enrichment)"

DEFAULT_DELAY_SECONDS = 1.5
DEFAULT_MAX_IMAGES = 5
MAX_RETRIES = 3
REQUEST_TIMEOUT = 30
MAX_BACKOFF_DELAY = 10
PLAYWRIGHT_TIMEOUT = 60000

SKU_COLUMN_CANDIDATES = [
    "sku", "SKU", "referencia", "Referencia",
    "modelo", "Modelo", "model", "Model",
]

IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
DOCUMENT_EXTENSIONS = {".pdf", ".doc", ".docx", ".txt"}

# Categorías conocidas de Hikvision (optimizadas para no navegar todas)
HIKVISION_CATEGORIES = [
    # Turbo HD Cameras (prioridad alta)
    "https://www.hikvision.com/es-co/products/Turbo-HD-Products/Turbo-HD-Cameras/ColorVu-Series/",
    "https://www.hikvision.com/es-co/products/Turbo-HD-Products/Turbo-HD-Cameras/Pro--Series/",
    "https://www.hikvision.com/es-co/products/Turbo-HD-Products/Turbo-HD-Cameras/Value-Series/",
    "https://www.hikvision.com/es-co/products/Turbo-HD-Products/Turbo-HD-Cameras/Ultra-Series/",
    "https://www.hikvision.com/es-co/products/Turbo-HD-Products/Turbo-HD-Cameras/IOT-Series/",
    # IP Network Cameras
    "https://www.hikvision.com/es-co/products/IP-Products/Network-Cameras/DeepinView-Series/",
    "https://www.hikvision.com/es-co/products/IP-Products/Network-Cameras/Pro-Series/",
]


# --- Logging ---

def setup_logging(verbose: bool = False) -> logging.Logger:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    return logging.getLogger("hikvision_enricher")


# --- Modelos de datos ---

@dataclass
class EnrichmentResult:
    sku_input: str
    sku_normalized: str
    sku_official: str = ""
    brand: str = "Hikvision"
    product_name: str = ""
    category_source: str = ""
    description: str = ""
    features_json: str = ""
    specifications_json: str = ""
    source_url: str = ""
    source_domain: str = ""
    datasheet_url: str = ""
    manual_url: str = ""
    quick_start_guide_url: str = ""
    additional_document_urls_json: str = ""
    image_source_urls_json: str = ""
    image_count: int = 0
    image_primary_path: str = ""
    image_local_paths_json: str = ""
    match_status: str = "pending"
    review_status: str = "pending"
    checked_at: str = ""
    notes: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "sku_input": self.sku_input,
            "sku_normalized": self.sku_normalized,
            "sku_official": self.sku_official,
            "brand": self.brand,
            "product_name": self.product_name,
            "category_source": self.category_source,
            "description": self.description,
            "features_json": self.features_json,
            "specifications_json": self.specifications_json,
            "source_url": self.source_url,
            "source_domain": self.source_domain,
            "datasheet_url": self.datasheet_url,
            "manual_url": self.manual_url,
            "quick_start_guide_url": self.quick_start_guide_url,
            "additional_document_urls_json": self.additional_document_urls_json,
            "image_source_urls_json": self.image_source_urls_json,
            "image_count": self.image_count,
            "image_primary_path": self.image_primary_path,
            "image_local_paths_json": self.image_local_paths_json,
            "match_status": self.match_status,
            "review_status": self.review_status,
            "checked_at": self.checked_at,
            "notes": self.notes,
        }


# --- Utilidades ---

def normalize_sku(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    s = re.sub(r"\s+", " ", s)
    s = s.upper()
    return s

def normalize_sku_compact(sku: str) -> str:
    s = normalize_sku(sku)
    s = s.replace("(", "").replace(")", "").replace(" ", "")
    return s.upper()

def is_valid_sku(sku: str) -> bool:
    return len(normalize_sku(sku)) > 0

def sanitize_filename(sku: str) -> str:
    s = normalize_sku(sku).replace(" ", "_").replace("-", "_")
    s = re.sub(r"[^A-Z0-9_]", "_", s)
    s = re.sub(r"_+", "_", s)
    return s.strip("_").upper()

def validate_domain(url: str) -> Optional[str]:
    try:
        parsed = urlparse(url)
        host = parsed.netloc.lower()
        if host in ALLOWED_DOMAINS:
            return host
        if host.endswith("hikvision.com"):
            return host
        return None
    except Exception:
        return None

def safe_json_dumps(obj: Any) -> str:
    try:
        return json.dumps(obj, ensure_ascii=False, default=str)
    except (TypeError, ValueError):
        return ""


def sku_in_html(html: str, target_sku: str) -> bool:
    """Verifica si el SKU está presente en cualquier parte del HTML (incluye JS)."""
    if not html:
        return False
    sku_clean = normalize_sku_compact(target_sku)
    # Buscar versión con y sin paréntesis
    variants = [
        target_sku.upper(),
        target_sku.upper().replace("(", "").replace(")", ""),
        sku_clean,
        normalize_sku(target_sku),
    ]
    html_upper = html.upper()
    for variant in variants:
        if variant in html_upper:
            return True
    return False


# --- Lectura de entrada ---

def detect_sku_column(columns: List[str]) -> Optional[str]:
    lower_cols = {c.lower(): c for c in columns}
    for candidate in SKU_COLUMN_CANDIDATES:
        if candidate.lower() in lower_cols:
            return lower_cols[candidate.lower()]
    return None

def read_input_file(input_path: str, sku_column: Optional[str]) -> "tuple[pd.DataFrame, str]":
    ext = Path(input_path).suffix.lower()
    if ext == ".csv":
        try:
            df = pd.read_csv(input_path, encoding="utf-8")
        except UnicodeDecodeError:
            df = pd.read_csv(input_path, encoding="utf-8-sig")
    elif ext in {".xlsx", ".xls"}:
        df = pd.read_excel(input_path, engine="openpyxl")
    else:
        raise ValueError(f"Formato de entrada no soportado: {ext}")

    columns = df.columns.tolist()
    if not sku_column:
        detected = detect_sku_column(columns)
        if not detected:
            raise ValueError(f"No se pudo detectar la columna de SKU. Use --sku-column. Columnas disponibles: {columns}")
        sku_column = detected

    return df, sku_column


# --- Navegador Playwright ---

def initialize_browser(logger: logging.Logger) -> Any:
    if sync_playwright is None:
        raise ImportError("Playwright no está instalado. Instale: pip install playwright && python -m playwright install chromium")

    try:
        playwright = sync_playwright().start()
        browser = playwright.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=USER_AGENT,
            viewport={"width": 1920, "height": 1080},
        )
        logger.info("Navegador Chromium iniciado (Playwright)")
        return browser, context
    except Exception as e:
        raise RuntimeError(f"Error inicializando Playwright: {e}")


def fetch_page_with_playwright(url: str, context: Any, logger: logging.Logger) -> Optional[BeautifulSoup]:
    try:
        page = context.new_page()
        page.goto(url, wait_until="networkidle", timeout=PLAYWRIGHT_TIMEOUT)
        time.sleep(3)
        html = page.content()
        page.close()
        return BeautifulSoup(html, "html.parser")
    except Exception as e:
        logger.debug(f"Error cargando {url} con Playwright: {e}")
        return None


def extract_product_urls_from_category(soup: BeautifulSoup) -> List[str]:
    product_urls = []
    seen = set()

    for a in soup.find_all("a", href=True):
        if not isinstance(a, Tag) or not isinstance(a["href"], str):
            continue
        href = str(a["href"])
        if "/products/" in href:
            full_url = urljoin("https://www.hikvision.com", href)
            if validate_domain(full_url):
                parts = href.strip("/").split("/")
                if len(parts) >= 5 and parts[0] == "es-co" and parts[1] == "products":
                    if full_url not in seen:
                        seen.add(full_url)
                        product_urls.append(full_url)

    return product_urls


def extract_json_ld(soup: BeautifulSoup) -> List[Dict[str, Any]]:
    scripts = soup.find_all("script", type="application/ld+json")
    results = []
    for script in scripts:
        if isinstance(script.string, str):
            try:
                data = json.loads(script.string)
                if isinstance(data, list):
                    results.extend(item for item in data if isinstance(item, dict))
                elif isinstance(data, dict):
                    results.append(data)
            except (json.JSONDecodeError, ValueError):
                continue
    return results

def extract_og_metadata(soup: BeautifulSoup) -> Dict[str, str]:
    meta_tags = {}
    for prop in ["og:title", "og:description", "og:image"]:
        tag = soup.find("meta", property=prop)
        if tag and isinstance(tag, Tag) and tag.get("content"):
            meta_tags[prop] = str(tag["content"])
    return meta_tags

def extract_image_urls(soup: BeautifulSoup) -> List[str]:
    urls = []
    seen = set()

    for img in soup.find_all("img"):
        if isinstance(img, Tag):
            src = img.get("src")
            if src and isinstance(src, str) and src not in seen:
                urls.append(src)
                seen.add(src)

    for img in soup.find_all("img"):
        if isinstance(img, Tag):
            srcset = img.get("srcset")
            if srcset and isinstance(srcset, str):
                for item in srcset.split(","):
                    url_part = item.strip().split(" ")[0]
                    if url_part and url_part not in seen:
                        urls.append(url_part)
                        seen.add(url_part)

    for link in soup.find_all("a", href=True):
        if isinstance(link, Tag) and isinstance(link["href"], str):
            href = str(link["href"])
            ext = Path(urlparse(href).path).suffix.lower()
            if ext in IMAGE_EXTENSIONS and href not in seen:
                urls.append(href)
                seen.add(href)

    return urls

def extract_document_urls(soup: BeautifulSoup) -> List[str]:
    urls = []
    seen = set()

    for link in soup.find_all("a", href=True):
        if isinstance(link, Tag) and isinstance(link["href"], str):
            href = str(link["href"])
            ext = Path(urlparse(href).path).suffix.lower()
            if ext in DOCUMENT_EXTENSIONS and href not in seen:
                full_url = urljoin("https://www.hikvision.com", href)
                if validate_domain(full_url):
                    urls.append(full_url)
                    seen.add(full_url)
    return urls

def extract_breadcrumb(soup: BeautifulSoup) -> str:
    breadcrumb_selectors = [
        "nav.breadcrumb", "div.breadcrumb", ".breadcrumb",
        "nav[aria-label='breadcrumb']", "ul.breadcrumb"
    ]
    for selector in breadcrumb_selectors:
        element = soup.select_one(selector)
        if element:
            texts = [el.get_text(strip=True) for el in element.find_all(["li", "span", "a"]) if el.get_text(strip=True)]
            if texts:
                return " > ".join(texts[1:-1]) if len(texts) > 2 else " > ".join(texts[1:])
    return ""

def extract_from_json_ld(json_ld: List[Dict[str, Any]], key: str, default: str = "") -> str:
    for item in json_ld:
        if key in item and isinstance(item[key], str):
            return item[key]
    return default

def extract_specifications(soup: BeautifulSoup, json_ld: List[Dict[str, Any]]) -> Dict[str, str]:
    """Extrae especificaciones técnicas desde JSON-LD (incluye additionalProperty) y HTML."""
    specs = {}

    for item in json_ld:
        if not isinstance(item, dict):
            continue

        item_type = item.get("@type", "")

        # Procesar Producto con additionalProperty
        if item_type == "Product" or item_type == "ProductModel":
            # additionalProperty contiene pares clave/valor técnicos
            additional_props = item.get("additionalProperty", [])
            if isinstance(additional_props, list):
                for prop in additional_props:
                    if isinstance(prop, dict):
                        name = prop.get("name", "")
                        value = prop.get("value", "")
                        if name and value:
                            if isinstance(value, str):
                                specs[name] = value
                            elif isinstance(value, (int, float, bool)):
                                specs[name] = str(value)

            # Extraer otros campos relevantes
            for k, v in item.items():
                if k not in {"@context", "@type", "name", "description", "sku", "image", "url",
                             "additionalProperty", "subjectOf", "offers", "aggregateRating", "brand", "model"}:
                    if isinstance(v, str) and v and not k.startswith("@"):
                        specs[k] = v

        # Continuar procesando todos los items JSON-LD
        # No hacer break - necesitamos procesar el segundo Product con additionalProperty

    # Si no hay especificaciones desde JSON-LD, intentar extraer desde HTML
    if not specs:
        # Buscar tablas de especificaciones
        for table in soup.find_all("table"):
            for row in table.find_all("tr"):
                cells = row.find_all(["td", "th"])
                if len(cells) >= 2:
                    key_elem = cells[0].get_text(strip=True)
                    value_elem = cells[1].get_text(strip=True)
                    if key_elem and value_elem:
                        specs[key_elem] = value_elem

        # Buscar divs con especificaciones técnicas
        if not specs:
            # Buscar patrones comunes de especificaciones
            spec_keywords = ["Sensor", "Resolución", "Lente", "Micrófono", "Iluminador",
                           "Día y noche", "Alimentación", "Dimensiones", "Peso", "Protección"]
            text = soup.get_text(" ", strip=True)

            for keyword in spec_keywords:
                idx = text.lower().find(keyword.lower())
                if idx != -1:
                    # Extraer contexto alrededor
                    context = text[idx:idx+200]
                    # Buscar patrón "Keyword: value"
                    next_idx = text.find(" ", idx + len(keyword))
                    if next_idx != -1:
                        value_start = next_idx + 1
                        value_end = min(value_start + 80, len(text))
                        value = text[value_start:value_end].strip()
                        if value and keyword not in specs:
                            specs[keyword] = value

    return specs


def classify_specs(specs: Dict[str, str]) -> Dict[str, Dict[str, str]]:
    """Clasifica especificaciones técnicas en secciones organizadas."""
    classified = {
        "Especificaciones": {},
        "Cámara": {},
        "Lente": {},
        "Micrófono": {},
        "Iluminador": {},
        "Imagen": {},
        "Interfaz": {},
        "General": {},
        "Aprobación": {},
    }

    # Mapeo de palabras clave a secciones
    section_mapping = {
        "sensor": "Cámara",
        "resolución": "Cámara",
        "iluminación": "Cámara",
        "obturador": "Cámara",
        "día y noche": "Cámara",
        "icr": "Cámara",
        "ángulo": "Cámara",
        "señal": "Cámara",
        "lente": "Lente",
        "focal": "Lente",
        "fov": "Lente",
        "montura": "Lente",
        "micrófono": "Micrófono",
        "captación": "Micrófono",
        "luz": "Iluminador",
        "infrarrojo": "Iluminador",
        "parámetro": "Imagen",
        "ajuste": "Imagen",
        "brillo": "Imagen",
        "nitidez": "Imagen",
        "frecuencia": "Imagen",
        "modo": "Imagen",
        "dinámico": "Imagen",
        "ruido": "Imagen",
        "balance": "Imagen",
        "mejora": "Imagen",
        "video": "Interfaz",
        "salida": "Interfaz",
        "material": "General",
        "alimentación": "General",
        "dimensiones": "General",
        "peso": "General",
        "funcionamiento": "General",
        "comunicación": "General",
        "idioma": "General",
        "protección": "Aprobación",
        "aprobación": "Aprobación",
        "ip": "Aprobación",
    }

    for key, value in specs.items():
        key_lower = key.lower()
        section = "Especificaciones"
        for keyword, sec in section_mapping.items():
            if keyword in key_lower:
                section = sec
                break

        if section not in classified:
            classified[section] = {}
        classified[section][key] = value

    return classified

def validate_sku_exact_match(soup: BeautifulSoup, target_sku: str, og_meta: Dict[str, str], json_ld: List[Dict[str, Any]]) -> bool:
    normalized_target = normalize_sku(target_sku)
    normalized_clean = normalized_target.replace("(", "").replace(")", "").replace(" ", "")
    normalized_compact = normalize_sku_compact(target_sku)
    sku_base = normalized_target.split("(")[0].strip().replace(" ", "").replace("-", "").upper()

    def sku_equivalent(sku_a: str, sku_b: str) -> bool:
        a = normalize_sku(sku_a).replace("(", "").replace(")", "").replace(" ", "").upper()
        b = normalize_sku(sku_b).replace("(", "").replace(")", "").replace(" ", "").upper()
        return a == b or a == normalize_sku_compact(sku_b)

    # 1. Title visible
    title_tag = soup.find("title")
    if title_tag and isinstance(title_tag, Tag) and title_tag.string:
        title_content = normalize_sku(title_tag.string)
        title_clean = title_content.replace("(", "").replace(")", "").replace(" ", "").upper()
        if (sku_equivalent(normalized_target, title_content) or
            sku_base in title_clean or
            normalized_compact in title_clean or
            normalized_clean in title_clean):
            return True

    # 2. og:title
    og_title = og_meta.get("og:title", "")
    if og_title:
        og_norm = normalize_sku(og_title)
        og_clean = og_norm.replace("(", "").replace(")", "").replace(" ", "").upper()
        if (sku_equivalent(normalized_target, og_title) or
            sku_base in og_clean or
            normalized_compact in og_clean or
            normalized_clean in og_clean):
            return True

    # 3. JSON-LD
    for item in json_ld:
        name = item.get("name", "")
        if name:
            name_norm = normalize_sku(name)
            name_clean = name_norm.replace("(", "").replace(")", "").replace(" ", "").upper()
            if (sku_equivalent(normalized_target, name) or
                sku_base in name_clean or
                normalized_clean in name_clean):
                return True
        sku_field = item.get("sku") or item.get("model")
        if sku_field:
            sku_norm = normalize_sku(sku_field)
            sku_clean = sku_norm.replace("(", "").replace(")", "").replace(" ", "").upper()
            if (sku_equivalent(normalized_target, sku_field) or
                normalized_compact == sku_clean or
                normalized_clean == sku_clean):
                return True

    # 4. Texto visible principal
    main_content = soup.find("main") or soup.find("body")
    if main_content:
        visible_text = main_content.get_text(" ", strip=True)
        visible_clean = re.sub(r"\s+", "", visible_text).upper()
        if (normalized_target in visible_text or
            normalized_clean in visible_clean or
            sku_base in visible_clean or
            normalized_compact in visible_clean):
            return True

    # 5. URL canónica y URL actual
    canonical = soup.find("link", {"rel": "canonical"})
    if canonical and isinstance(canonical, Tag) and canonical.get("href"):
        url_href = str(canonical["href"])
        url_upper = url_href.upper()
        url_clean = re.sub(r"[^\w]", "", url_upper)
        if (normalized_clean in url_upper or
            sku_base in url_upper or
            normalized_compact in url_clean):
            return True

    return False


def validate_sku_partial_match(soup: BeautifulSoup, target_sku: str) -> bool:
    normalized_target = normalize_sku(target_sku)
    normalized_clean = normalized_target.replace("(", "").replace(")", "").replace(" ", "")
    sku_base = normalized_target.split("(")[0].strip().replace(" ", "").replace("-", "").upper()

    title_tag = soup.find("title")
    title_text = ""
    if title_tag and isinstance(title_tag, Tag) and title_tag.string:
        title_text = normalize_sku(title_tag.string)

    main_content = soup.find("main") or soup.find("body")
    visible_text = main_content.get_text(" ", strip=True) if main_content else ""
    visible_clean = visible_text.upper()

    if sku_base in title_text.upper() or sku_base in visible_clean:
        if normalized_clean not in title_text.upper().replace("(", "").replace(")", "").replace(" ", "") and \
           normalized_clean not in visible_clean.replace("(", "").replace(")", "").replace(" ", ""):
            return True

    return False


def download_document(url: str, local_path: Path, logger: logging.Logger) -> bool:
    """Descarga un documento (PDF, manual, etc.)."""
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return False
        content_type = resp.headers.get("Content-Type", "")
        if not (content_type.startswith("application/") or content_type.startswith("text/")):
            logger.debug(f"Rechazado documento no válido ({content_type}): {url}")
            return False
        local_path.write_bytes(resp.content)
        return True
    except Exception as e:
        logger.debug(f"Error descargando documento {url}: {e}")
        return False


def download_image(url: str, local_path: Path, logger: logging.Logger) -> bool:
    try:
        resp = requests.get(url, headers={"User-Agent": USER_AGENT}, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return False
        content_type = resp.headers.get("Content-Type", "")
        if not content_type.startswith("image/"):
            logger.debug(f"Rechazada imagen no válida ({content_type}): {url}")
            return False
        local_path.write_bytes(resp.content)
        return True
    except Exception as e:
        logger.debug(f"Error descargando imagen {url}: {e}")
        return False


def extract_assets_from_json_ld(json_ld: List[Dict[str, Any]]) -> Dict[str, Any]:
    """Extrae URLs de imágenes y documentos desde JSON-LD."""
    assets = {
        "images": [],
        "documents": [],
        "brand": "",
        "description": "",
        "product_name": "",
        "sku": "",
    }

    for item in json_ld:
        if not isinstance(item, dict):
            continue

        item_type = item.get("@type", "")

        # ImageObject: contiene URLs de imágenes principales
        if item_type == "ImageObject":
            content_url = item.get("contentUrl")
            if content_url and isinstance(content_url, str):
                if content_url not in assets["images"]:
                    assets["images"].append(content_url)
            continue

        # Product: extraer nombre, descripción, marca, SKU, imágenes, documentos
        if item_type == "Product":
            if "name" in item and isinstance(item["name"], str):
                if not assets["product_name"]:
                    assets["product_name"] = item["name"]
            if "description" in item and isinstance(item["description"], str):
                if not assets["description"]:
                    assets["description"] = item["description"]
            if "brand" in item:
                brand = item["brand"]
                if isinstance(brand, str):
                    assets["brand"] = brand
                elif isinstance(brand, dict) and isinstance(brand.get("name"), str):
                    assets["brand"] = brand["name"]
            if "model" in item and isinstance(item["model"], str):
                if not assets["sku"]:
                    assets["sku"] = item["model"]
            elif "sku" in item and isinstance(item["sku"], str):
                if not assets["sku"]:
                    assets["sku"] = item["sku"]

            # Extraer imágenes del producto (image puede ser string o lista)
            image_field = item.get("image")
            if isinstance(image_field, str):
                if image_field not in assets["images"]:
                    assets["images"].append(image_field)
            elif isinstance(image_field, list):
                for img in image_field:
                    if isinstance(img, str) and img not in assets["images"]:
                        assets["images"].append(img)

            # Extraer documentos desde subjectOf (estructura Hikvision)
            subject_of = item.get("subjectOf", [])
            if isinstance(subject_of, dict):
                subject_of = [subject_of]
            if isinstance(subject_of, list):
                for part in subject_of:
                    if not isinstance(part, dict):
                        continue
                    url_val = part.get("url") or part.get("contentUrl", "")
                    if url_val and isinstance(url_val, str):
                        path = urlparse(url_val).path.lower()
                        if any(path.endswith(ext) for ext in DOCUMENT_EXTENSIONS):
                            if url_val not in assets["documents"]:
                                assets["documents"].append(url_val)
                        elif any(path.endswith(ext) for ext in IMAGE_EXTENSIONS):
                            if url_val not in assets["images"]:
                                assets["images"].append(url_val)

            # Extraer URLs de cualquier campo que parezca URL
            for key, value in item.items():
                if isinstance(value, str) and "http" in value:
                    path = urlparse(value).path.lower()
                    if any(path.endswith(ext) for ext in IMAGE_EXTENSIONS):
                        if value not in assets["images"]:
                            assets["images"].append(value)
                    elif any(path.endswith(ext) for ext in DOCUMENT_EXTENSIONS):
                        if value not in assets["documents"]:
                            assets["documents"].append(value)

    return assets


# --- Procesamiento principal ---

def process_sku(sku_input: str, output_dir: str, max_images: int, delay_seconds: float,
                dry_run: bool, logger: logging.Logger, context: Any,
                source_map: Dict[str, str],
                category_cache: Dict[str, List[str]]) -> EnrichmentResult:
    """Procesa un SKU individual y retorna el resultado enriquecido."""
    result = EnrichmentResult(
        sku_input=normalize_sku(sku_input) if sku_input else "",
        sku_normalized=normalize_sku(sku_input) if sku_input else "",
        checked_at=datetime.now(timezone.utc).isoformat(),
    )

    if not sku_input or not is_valid_sku(sku_input):
        result.match_status = "invalid_sku"
        result.notes = "SKU vacío o inválido."
        return result

    normalized_target = result.sku_normalized
    normalized_clean = normalized_target.replace("(", "").replace(")", "")
    normalized_compact = normalize_sku_compact(normalized_target)
    sku_base = normalized_target.split("(")[0].strip().replace(" ", "").replace("-", "").upper()

    # Priorizar source_map si existe
    product_url = source_map.get(normalized_target) or source_map.get(normalized_clean)

    if product_url:
        logger.info(f"Usando URL del source-map para {normalized_target}: {product_url}")
        product_urls = [product_url]
    else:
        # Estrategia: navegar categorías y buscar URLs de productos que contengan el SKU base
        logger.debug(f"Buscando en categorías por SKU: {normalized_target}")
        product_urls = []

        for category_url in HIKVISION_CATEGORIES:
            try:
                # Usar caché de categorías si está disponible
                if category_url in category_cache:
                    category_product_urls = category_cache[category_url]
                    logger.debug(f"Categoría (cache): {category_url} -> {len(category_product_urls)} productos")
                else:
                    soup = fetch_page_with_playwright(category_url, context, logger)
                    if not soup:
                        continue
                    category_product_urls = extract_product_urls_from_category(soup)
                    category_cache[category_url] = category_product_urls
                    logger.debug(f"Categoría {category_url}: {len(category_product_urls)} productos encontrados")

                # Filtrar URLs que contengan el SKU base en la ruta
                for url in category_product_urls:
                    url_compact = re.sub(r"[^\w]", "", url.upper())
                    if sku_base in url_compact or normalized_clean in url.upper().replace("(", "").replace(")", ""):
                        if url not in product_urls:
                            product_urls.append(url)

                # Si ya tenemos suficientes candidatos, podemos detenernos
                if len(product_urls) >= 5:
                    break

            except Exception as e:
                logger.debug(f"Error navegando categoría {category_url}: {e}")
                continue

        # Si no encontramos URLs, intentar navegar directamente a la URL del producto
        if not product_urls:
            direct_url = f"https://www.hikvision.com/es-co/products/turbo-hd-products/turbo-hd-cameras/colorvu-series/{sku_base.lower()}/"
            product_urls = [direct_url]

    if not product_urls:
        result.match_status = "not_found"
        result.notes = "No se encontraron URLs de producto en resultados de navegación."
        return result

    logger.debug(f"Validando {len(product_urls)} URLs de productos")

    exact_match_url = None
    exact_match_soup = None
    partial_match_urls = []

    for product_url in product_urls[:10]:
        try:
            soup = fetch_page_with_playwright(product_url, context, logger)
            if not soup:
                continue

            # Verificar dominio
            if not validate_domain(product_url):
                logger.debug(f"Dominio no permitido: {product_url}")
                continue

            og_meta = extract_og_metadata(soup)
            json_ld_data = extract_json_ld(soup)

            # Validar SKU exacto (métodos estructurados + HTML completo)
            og_meta = extract_og_metadata(soup)
            json_ld_data = extract_json_ld(soup)

            if validate_sku_exact_match(soup, normalized_target, og_meta, json_ld_data):
                exact_match_url = product_url
                exact_match_soup = soup
                logger.debug(f"Coincidencia exacta encontrada: {product_url}")
                break

            # Validación adicional: buscar SKU en HTML completo (incluye JS embebido)
            # Esto es útil cuando el SKU está en elementos dinámicos no capturados por métodos anteriores
            html_content = soup.prettify() if soup else ""
            if sku_in_html(html_content, normalized_target):
                exact_match_url = product_url
                exact_match_soup = soup
                logger.debug(f"Coincidencia exacta (HTML completo) encontrada: {product_url}")
                break

            # Validar coincidencia parcial
            if validate_sku_partial_match(soup, normalized_target):
                partial_match_urls.append(product_url)

        except Exception as e:
            logger.debug(f"Error validando producto {product_url}: {e}")
            continue

    if exact_match_url:
        # Reutilizar el soup de la validación si está disponible
        try:
            if 'exact_match_soup' in locals() and exact_match_soup:
                soup = exact_match_soup
            else:
                soup = fetch_page_with_playwright(exact_match_url, context, logger)
            if not soup:
                result.match_status = "error"
                result.notes = "Error cargando página del producto exacto."
                return result
        except Exception as e:
            logger.debug(f"Error cargando página exacta: {e}")
            result.match_status = "error"
            result.notes = f"Error cargando página del producto: {e}"
            return result

        og_meta = extract_og_metadata(soup)
        json_ld_data = extract_json_ld(soup)
        json_ld_assets = extract_assets_from_json_ld(json_ld_data)

        result.source_url = exact_match_url
        result.source_domain = urlparse(exact_match_url).netloc.lower()

        # Priorizar datos del JSON-LD (más estructurados) sobre og_meta
        result.product_name = json_ld_assets.get("product_name") or og_meta.get("og:title", "") or extract_from_json_ld(json_ld_data, "name", "")
        result.description = json_ld_assets.get("description") or og_meta.get("og:description", "") or extract_from_json_ld(json_ld_data, "description", "")
        result.category_source = extract_breadcrumb(soup)

        specs = extract_specifications(soup, json_ld_data)
        # Añadir SKU oficial si está disponible
        if json_ld_assets.get("sku"):
            specs["sku"] = json_ld_assets["sku"]
        result.specifications_json = safe_json_dumps(specs)

        # Extraer assets desde JSON-LD y HTML
        image_urls = extract_image_urls(soup)
        doc_urls = extract_document_urls(soup)

        # Añadir imágenes desde JSON-LD
        for img_url in json_ld_assets.get("images", []):
            if img_url and isinstance(img_url, str):
                full_url = urljoin(exact_match_url, img_url)
                if validate_domain(full_url) and full_url not in image_urls:
                    image_urls.append(full_url)

        # Añadir documentos desde JSON-LD
        for doc_url in json_ld_assets.get("documents", []):
            if doc_url and isinstance(doc_url, str):
                full_url = urljoin(exact_match_url, doc_url)
                if validate_domain(full_url) and full_url not in doc_urls:
                    doc_urls.append(full_url)

        logger.debug(f"Total imágenes (HTML + JSON-LD): {len(image_urls)}")
        logger.debug(f"Total documentos (HTML + JSON-LD): {len(doc_urls)}")

        unique_images: set[str] = set()
        large_images: List[str] = []
        thumb_images: List[str] = []

        for url in image_urls:
            ext = Path(urlparse(url).path).suffix.lower()
            if ext not in IMAGE_EXTENSIONS:
                continue  # Saltar imágenes sin extensión válida
            full_url = urljoin(exact_match_url, url)
            if validate_domain(full_url) and full_url not in unique_images:
                unique_images.add(full_url)
                # Separar imágenes grandes de miniaturas
                if ".thumb." in full_url or "/specicon/" in full_url or "/icons/" in full_url:
                    thumb_images.append(full_url)
                else:
                    large_images.append(full_url)

        # Priorizar imágenes grandes sobre miniaturas e iconos
        # Solo usar imágenes grandes si hay disponibles
        if large_images:
            valid_images = large_images[:max_images]
        else:
            valid_images = (large_images + thumb_images)[:max_images]

        logger.debug(f"Valid images: {len(valid_images)} (large: {len(large_images)}, thumbs: {len(thumb_images)}")

        result.image_source_urls_json = safe_json_dumps(valid_images)
        result.image_count = len(valid_images)

        unique_docs = set()
        filtered_docs = []
        for url in doc_urls:
            if url not in unique_docs and validate_domain(url):
                unique_docs.add(url)
                filtered_docs.append(url)

        for doc_url in filtered_docs:
            path = urlparse(doc_url).path.lower()
            if "datasheet" in path or "ficha" in path:
                result.datasheet_url = doc_url
            elif "manual" in path:
                result.manual_url = doc_url
            elif "quick" in path or "guia" in path or "start" in path:
                result.quick_start_guide_url = doc_url
            else:
                docs_list = json.loads(result.additional_document_urls_json) if result.additional_document_urls_json else []
                docs_list.append(doc_url)
                result.additional_document_urls_json = safe_json_dumps(docs_list)

        # Clasificar especificaciones técnicas
        specs = extract_specifications(soup, json_ld_data)
        if json_ld_assets.get("sku"):
            specs["sku"] = json_ld_assets["sku"]
        classified_specs = classify_specs(specs)
        result.specifications_json = safe_json_dumps(classified_specs)

        if result.product_name and result.specifications_json and valid_images:
            result.match_status = "matched"
        elif result.product_name or result.specifications_json or result.image_count > 0:
            result.match_status = "partial"
        else:
            result.match_status = "matched"

        # Descargar imágenes y documentos en carpetas por SKU
        if not dry_run and result.match_status in {"matched", "partial"}:
            safe_sku = sanitize_filename(normalized_target)
            # Estructura: output_dir/<sku>/images/ y output_dir/<sku>/pdfs/
            sku_dir = Path(output_dir) / safe_sku
            images_dir = sku_dir / "images"
            images_dir.mkdir(parents=True, exist_ok=True)
            pdfs_dir = sku_dir / "pdfs"
            pdfs_dir.mkdir(parents=True, exist_ok=True)

            local_paths = []

            for idx, url in enumerate(valid_images, start=1):
                ext = Path(urlparse(url).path).suffix.lower() or ".jpg"
                if ext not in IMAGE_EXTENSIONS:
                    ext = ".jpg"
                filename = f"hikvision__{safe_sku}__{idx:02d}{ext}"
                local_path = images_dir / filename

                if local_path.exists() and local_path.stat().st_size > 0:
                    local_paths.append(str(local_path.relative_to(output_dir)))
                    continue

                if download_image(url, local_path, logger):
                    local_paths.append(str(local_path.relative_to(output_dir)))
                    logger.debug(f"Imagen guardada: {local_path}")

            result.image_local_paths_json = safe_json_dumps(local_paths)
            if local_paths:
                result.image_primary_path = local_paths[0]

            # Descargar documentos (datasheet, manuales)
            downloaded_docs = {}
            for doc_field, doc_url in [("datasheet", result.datasheet_url), ("manual", result.manual_url), ("quick_start", result.quick_start_guide_url)]:
                if doc_url and not dry_run:
                    try:
                        doc_ext = Path(urlparse(doc_url).path).suffix.lower() or ".pdf"
                        if doc_ext not in DOCUMENT_EXTENSIONS:
                            doc_ext = ".pdf"
                        doc_filename = f"hikvision__{safe_sku}__{doc_field}{doc_ext}"
                        doc_path = pdfs_dir / doc_filename

                        if doc_path.exists() and doc_path.stat().st_size > 0:
                            downloaded_docs[doc_field] = str(doc_path.relative_to(output_dir))
                            continue

                        if download_document(doc_url, doc_path, logger):
                            downloaded_docs[doc_field] = str(doc_path.relative_to(output_dir))
                            logger.debug(f"Documento guardado: {doc_path}")
                    except Exception as e:
                        logger.debug(f"Error descargando {doc_field}: {e}")

            if downloaded_docs:
                result.notes = f"Documentos descargados: {', '.join(downloaded_docs.keys())}"

        return result

    elif partial_match_urls:
        result.match_status = "ambiguous"
        result.notes = f"Se encontraron {len(partial_match_urls)} coincidencias parciales. URLs: {', '.join(partial_match_urls[:3])}"
        return result

    result.match_status = "not_found"
    result.notes = "SKU no encontrado en páginas de producto individuales."
    return result


def load_existing_results(output_dir: str) -> Dict[str, str]:
    resume_map = {}
    log_path = Path(output_dir) / "enrichment_log.csv"
    if log_path.exists():
        with log_path.open("r", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for row in reader:
                status = row.get("estado", "")
                if status in {"matched", "partial", "ambiguous", "not_found"}:
                    resume_map[row["sku"]] = status
    return resume_map

def load_source_map(source_map_path: str) -> Dict[str, str]:
    mapping = {}
    if not source_map_path:
        return mapping
    df = pd.read_csv(source_map_path)
    for _, row in df.iterrows():
        sku = normalize_sku(row.get("sku", ""))
        url = str(row.get("source_url", "")).strip()
        if sku and url:
            mapping[sku] = url
            mapping[sku.replace("(", "").replace(")", "").strip()] = url
    return mapping

def save_outputs(results: List[EnrichmentResult], df_original: pd.DataFrame,
                 sku_column: str, output_dir: str, logger: logging.Logger) -> None:
    """Genera un solo archivo Excel con todas las hojas necesarias."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_path / "catalogo_hikvision_enriquecido.xlsx"

    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage

        # Eliminar archivo existente
        if xlsx_path.exists():
            try:
                xlsx_path.unlink()
            except:
                pass

        wb = openpyxl.Workbook()

        # === HOJA 1: Datos ===
        ws_data = wb.active
        ws_data.title = "Datos"

        # Headers
        headers = ["SKU", "Estado", "Nombre Producto", "Descripción", "URL Fuente", "Datasheet PDF", "Manual PDF"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        for row_idx, result in enumerate(results, 2):
            ws_data.cell(row=row_idx, column=1, value=result.sku_input)
            ws_data.cell(row=row_idx, column=2, value=result.match_status)
            ws_data.cell(row=row_idx, column=3, value=result.product_name)
            ws_data.cell(row=row_idx, column=4, value=result.description)
            ws_data.cell(row=row_idx, column=5, value=result.source_url)

            # Referenciar PDFs relativos
            safe_sku = sanitize_filename(result.sku_normalized)
            datasheet_path = output_path / safe_sku / "pdfs" / f"hikvision__{safe_sku}__datasheet.pdf"
            manual_path = output_path / safe_sku / "pdfs" / f"hikvision__{safe_sku}__manual.pdf"

            ws_data.cell(row=row_idx, column=6, value=str(datasheet_path.relative_to(output_path)) if datasheet_path.exists() else result.datasheet_url)
            ws_data.cell(row=row_idx, column=7, value=str(manual_path.relative_to(output_path)) if manual_path.exists() else result.manual_url)

        ws_data.column_dimensions["A"].width = 25
        ws_data.column_dimensions["B"].width = 12
        ws_data.column_dimensions["C"].width = 40
        ws_data.column_dimensions["D"].width = 60
        ws_data.column_dimensions["E"].width = 50
        ws_data.column_dimensions["F"].width = 40
        ws_data.column_dimensions["G"].width = 40

        for row in range(2, ws_data.max_row + 1):
            ws_data.row_dimensions[row].height = 20

        # === HOJA 2: Especificaciones ===
        ws_specs = wb.create_sheet("Especificaciones")
        spec_headers = ["SKU", "Sección", "Campo", "Valor"]
        for col_idx, header in enumerate(spec_headers, 1):
            cell = ws_specs.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        spec_row = 2
        spec_sections = ["Especificaciones", "Cámara", "Lente", "Micrófono", "Iluminador",
                        "Imagen", "Interfaz", "General", "Aprobación"]

        for result in results:
            if result.match_status in {"matched", "partial"} and result.specifications_json:
                try:
                    classified = json.loads(result.specifications_json)
                    for section in spec_sections:
                        if section in classified and classified[section]:
                            for field, value in classified[section].items():
                                if field.startswith("@") or field.startswith("http"):
                                    continue
                                if not field or not value:
                                    continue
                                ws_specs.cell(row=spec_row, column=1, value=result.sku_input)
                                ws_specs.cell(row=spec_row, column=2, value=section)
                                ws_specs.cell(row=spec_row, column=3, value=field)
                                ws_specs.cell(row=spec_row, column=4, value=value)

                                category_colors = {
                                    "Cámara": "E2EFDA", "Lente": "FCE4D6", "Micrófono": "E2EFDA",
                                    "Iluminador": "FFF2CC", "Imagen": "E1F3DB",
                                    "Interfaz": "D9E1F2", "General": "D9D2E9",
                                    "Aprobación": "FCE4D6", "Especificaciones": "FFFFFF",
                                }
                                fill_color = category_colors.get(section, "FFFFFF")
                                for col in range(1, 5):
                                    ws_specs.cell(row=spec_row, column=col).fill = openpyxl.styles.PatternFill(
                                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                                    )
                                spec_row += 1
                except (json.JSONDecodeError, TypeError):
                    pass

        ws_specs.column_dimensions["A"].width = 25
        ws_specs.column_dimensions["B"].width = 20
        ws_specs.column_dimensions["C"].width = 35
        ws_specs.column_dimensions["D"].width = 60

        for row in range(2, ws_specs.max_row + 1):
            ws_specs.row_dimensions[row].height = 20

        # === HOJA 3: Imágenes ===
        ws_images = wb.create_sheet("Imágenes")
        ws_images.cell(row=1, column=1, value="SKU")
        ws_images.cell(row=1, column=2, value="Estado")
        ws_images.cell(row=1, column=3, value="Imagen")

        for col in range(1, 4):
            cell = ws_images.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        img_row = 2
        for result in results:
            if result.image_local_paths_json and result.match_status in {"matched", "partial"}:
                try:
                    local_paths = json.loads(result.image_local_paths_json)
                    if local_paths:
                        img_path = Path(output_dir) / local_paths[0]
                        if img_path.exists() and img_path.stat().st_size > 1000:
                            ws_images.cell(row=img_row, column=1, value=result.sku_input)
                            ws_images.cell(row=img_row, column=2, value=result.match_status)
                            img = XLImage(str(img_path))
                            img.width = 400
                            img.height = 400
                            ws_images.add_image(img, f"C{img_row}")
                            ws_images.row_dimensions[img_row].height = 300
                            ws_images.column_dimensions["C"].width = 50
                            img_row += 1
                except Exception as e:
                    logger.debug(f"Error insertando imagen: {e}")

        if img_row == 2:
            ws_images.cell(row=2, column=1, value="No hay imágenes para mostrar")

        wb.save(xlsx_path)
        logger.info(f"XLSX guardado: {xlsx_path}")
    except Exception as e:
        logger.error(f"Error guardando XLSX: {e}")




def load_source_map(source_map_path: str) -> Dict[str, str]:
    mapping = {}
    if not source_map_path:
        return mapping
    df = pd.read_csv(source_map_path)
    for _, row in df.iterrows():
        sku = normalize_sku(row.get("sku", ""))
        url = str(row.get("source_url", "")).strip()
        if sku and url:
            mapping[sku] = url
            mapping[sku.replace("(", "").replace(")", "").strip()] = url
    return mapping

def save_outputs(results: List[EnrichmentResult], df_original: pd.DataFrame,
                 sku_column: str, output_dir: str, logger: logging.Logger) -> None:
    """Genera un solo archivo Excel con todas las hojas necesarias."""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    xlsx_path = output_path / "catalogo_hikvision_enriquecido.xlsx"

    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage

        # Eliminar archivo existente
        if xlsx_path.exists():
            try:
                xlsx_path.unlink()
            except:
                pass

        wb = openpyxl.Workbook()

        # === HOJA 1: Datos ===
        ws_data = wb.active
        ws_data.title = "Datos"

        # Headers
        headers = ["SKU", "Estado", "Nombre Producto", "Descripción", "URL Fuente", "Datasheet PDF", "Manual PDF"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_data.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        for row_idx, result in enumerate(results, 2):
            ws_data.cell(row=row_idx, column=1, value=result.sku_input)
            ws_data.cell(row=row_idx, column=2, value=result.match_status)
            ws_data.cell(row=row_idx, column=3, value=result.product_name)
            ws_data.cell(row=row_idx, column=4, value=result.description)
            ws_data.cell(row=row_idx, column=5, value=result.source_url)

            # Referenciar PDFs relativos
            safe_sku = sanitize_filename(result.sku_normalized)
            datasheet_path = output_path / safe_sku / "pdfs" / f"hikvision__{safe_sku}__datasheet.pdf"
            manual_path = output_path / safe_sku / "pdfs" / f"hikvision__{safe_sku}__manual.pdf"

            ws_data.cell(row=row_idx, column=6, value=str(datasheet_path.relative_to(output_path)) if datasheet_path.exists() else result.datasheet_url)
            ws_data.cell(row=row_idx, column=7, value=str(manual_path.relative_to(output_path)) if manual_path.exists() else result.manual_url)

        ws_data.column_dimensions["A"].width = 25
        ws_data.column_dimensions["B"].width = 12
        ws_data.column_dimensions["C"].width = 40
        ws_data.column_dimensions["D"].width = 60
        ws_data.column_dimensions["E"].width = 50
        ws_data.column_dimensions["F"].width = 40
        ws_data.column_dimensions["G"].width = 40

        for row in range(2, ws_data.max_row + 1):
            ws_data.row_dimensions[row].height = 20

        # === HOJA 2: Especificaciones ===
        ws_specs = wb.create_sheet("Especificaciones")
        spec_headers = ["SKU", "Sección", "Campo", "Valor"]
        for col_idx, header in enumerate(spec_headers, 1):
            cell = ws_specs.cell(row=1, column=col_idx, value=header)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        spec_row = 2
        spec_sections = ["Especificaciones", "Cámara", "Lente", "Micrófono", "Iluminador",
                        "Imagen", "Interfaz", "General", "Aprobación"]

        for result in results:
            if result.match_status in {"matched", "partial"} and result.specifications_json:
                try:
                    classified = json.loads(result.specifications_json)
                    for section in spec_sections:
                        if section in classified and classified[section]:
                            for field, value in classified[section].items():
                                if field.startswith("@") or field.startswith("http"):
                                    continue
                                if not field or not value:
                                    continue
                                ws_specs.cell(row=spec_row, column=1, value=result.sku_input)
                                ws_specs.cell(row=spec_row, column=2, value=section)
                                ws_specs.cell(row=spec_row, column=3, value=field)
                                ws_specs.cell(row=spec_row, column=4, value=value)

                                category_colors = {
                                    "Cámara": "E2EFDA", "Lente": "FCE4D6", "Micrófono": "E2EFDA",
                                    "Iluminador": "FFF2CC", "Imagen": "E1F3DB",
                                    "Interfaz": "D9E1F2", "General": "D9D2E9",
                                    "Aprobación": "FCE4D6", "Especificaciones": "FFFFFF",
                                }
                                fill_color = category_colors.get(section, "FFFFFF")
                                for col in range(1, 5):
                                    ws_specs.cell(row=spec_row, column=col).fill = openpyxl.styles.PatternFill(
                                        start_color=fill_color, end_color=fill_color, fill_type="solid"
                                    )
                                spec_row += 1
                except (json.JSONDecodeError, TypeError):
                    pass

        ws_specs.column_dimensions["A"].width = 25
        ws_specs.column_dimensions["B"].width = 20
        ws_specs.column_dimensions["C"].width = 35
        ws_specs.column_dimensions["D"].width = 60

        for row in range(2, ws_specs.max_row + 1):
            ws_specs.row_dimensions[row].height = 20

        # === HOJA 3: Imágenes ===
        ws_images = wb.create_sheet("Imágenes")
        ws_images.cell(row=1, column=1, value="SKU")
        ws_images.cell(row=1, column=2, value="Estado")
        ws_images.cell(row=1, column=3, value="Imagen")

        for col in range(1, 4):
            cell = ws_images.cell(row=1, column=col)
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(start_color="366092", fill_type="solid")

        img_row = 2
        for result in results:
            if result.image_local_paths_json and result.match_status in {"matched", "partial"}:
                try:
                    local_paths = json.loads(result.image_local_paths_json)
                    if local_paths:
                        img_path = Path(output_dir) / local_paths[0]
                        if img_path.exists() and img_path.stat().st_size > 1000:
                            ws_images.cell(row=img_row, column=1, value=result.sku_input)
                            ws_images.cell(row=img_row, column=2, value=result.match_status)
                            img = XLImage(str(img_path))
                            img.width = 400
                            img.height = 400
                            ws_images.add_image(img, f"C{img_row}")
                            ws_images.row_dimensions[img_row].height = 300
                            ws_images.column_dimensions["C"].width = 50
                            img_row += 1
                except Exception as e:
                    logger.debug(f"Error insertando imagen: {e}")

        if img_row == 2:
            ws_images.cell(row=2, column=1, value="No hay imágenes para mostrar")

        wb.save(xlsx_path)
        logger.info(f"XLSX guardado: {xlsx_path}")
    except Exception as e:
        logger.error(f"Error guardando XLSX: {e}")
        try:
            if xlsx_path.exists():
                try:
                    xlsx_path.unlink()
                except:
                    pass
            basic_data.to_excel(xlsx_path, index=False, engine="openpyxl")
            logger.info(f"XLSX guardado (formato básico): {xlsx_path}")
        except Exception as e2:
            logger.error(f"Error crítico guardando XLSX: {e2}")

    # JSON
    json_path = output_path / "catalogo_hikvision_especificaciones.json"
    specs_list = [{
        "sku_input": r.sku_input,
        "match_status": r.match_status,
        "product_name": r.product_name,
        "specifications_json": r.specifications_json,
        "source_url": r.source_url,
        "image_local_paths_json": r.image_local_paths_json,
        "checked_at": r.checked_at,
    } for r in results]
    json_path.write_text(safe_json_dumps(specs_list), encoding="utf-8")

    # Log
    log_path = output_path / "enrichment_log.csv"
    with log_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["sku", "estado", "url_final", "imagenes_encontradas",
                         "imagenes_descargadas", "hora", "mensaje"])
        for r in results:
            downloaded = len(json.loads(r.image_local_paths_json)) if r.image_local_paths_json else 0
            writer.writerow([r.sku_input, r.match_status, r.source_url, r.image_count,
                            downloaded, r.checked_at, r.notes])


class HikvisionEnricher:
    def __init__(self, args: argparse.Namespace, logger: logging.Logger):
        self.args = args
        self.logger = logger
        self.context: Any = None
        self.browser: Any = None
        self.category_cache: Dict[str, List[str]] = {}

    def run(self) -> int:
        try:
            df, sku_column = read_input_file(self.args.input, self.args.sku_column)
        except Exception as e:
            self.logger.error(f"Error al leer archivo de entrada: {e}")
            return 1

        total_rows = len(df)
        results: List[EnrichmentResult] = []

        resume_map: Dict[str, str] = {}
        if self.args.resume:
            resume_map = load_existing_results(self.args.output_dir)

        source_map: Dict[str, str] = {}
        if self.args.source_map:
            try:
                source_map = load_source_map(self.args.source_map)
                self.logger.info(f"Cargados {len(source_map)} mapeos del source-map")
            except Exception as e:
                self.logger.error(f"Error cargando source-map: {e}")
                return 1

        rows_to_process = df.head(self.args.limit) if self.args.limit else df

        try:
            self.browser, self.context = initialize_browser(self.logger)
            self.logger.info("Navegador listo para procesamiento")
        except Exception as e:
            self.logger.error(f"Error inicializando navegador: {e}")
            return 1

        try:
            for idx, row in rows_to_process.iterrows():
                if idx > 0:
                    time.sleep(self.args.delay_seconds)

                sku_value = row[sku_column]
                sku_input = str(sku_value).strip() if pd.notna(sku_value) else ""

                normalized = normalize_sku(sku_input)
                if normalized in resume_map:
                    self.logger.info(f"Saltando (ya procesado): {normalized} -> {resume_map[normalized]}")
                    continue

                self.logger.info(f"[{idx+1}/{len(rows_to_process)}] Procesando SKU: {normalized or '(inválido)'}")
                result = process_sku(
                    sku_input,
                    self.args.output_dir,
                    self.args.max_images,
                    self.args.delay_seconds,
                    self.args.dry_run,
                    self.logger,
                    self.context,
                    source_map,
                    self.category_cache,
                )
                results.append(result)

            if not results:
                self.logger.warning("No se procesaron registros nuevos.")
                return 0

            save_outputs(results, df, sku_column, self.args.output_dir, self.logger)
            self.print_summary(results, total_rows)
            return 0
        finally:
            if self.browser:
                self.browser.close()
                self.logger.info("Navegador cerrado")

    def print_summary(self, results: List[EnrichmentResult], total_rows: int) -> None:
        from collections import Counter
        counts = Counter(r.match_status for r in results)
        matched = counts.get("matched", 0)
        partial = counts.get("partial", 0)
        ambiguous = counts.get("ambiguous", 0)
        not_found = counts.get("not_found", 0)
        invalid_sku = counts.get("invalid_sku", 0)
        error = counts.get("error", 0)
        images_downloaded = sum(
            len(json.loads(r.image_local_paths_json)) if r.image_local_paths_json else 0
            for r in results
        )

        print("\n=== RESUMEN ENRIQUECIMIENTO HIKVISION ===")
        print(f"Total filas analizadas: {total_rows}")
        print(f"Matched:                 {matched}")
        print(f"Partial:                 {partial}")
        print(f"Ambiguous:               {ambiguous}")
        print(f"Not found:               {not_found}")
        print(f"Invalid SKU:             {invalid_sku}")
        print(f"Error:                   {error}")
        print(f"Imagenes descargadas:    {images_downloaded}")
        print(f"Directorio de salida:    {Path(self.args.output_dir).resolve()}")


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        prog="enrich_hikvision_catalog",
        description="Enriquecedor de catálogo Hikvision por SKU — navegador real (Playwright).",
    )
    parser.add_argument("--input", "-i", required=True,
                        help="Archivo de entrada (.csv o .xlsx) con lista de SKU.")
    parser.add_argument("--output-dir", "-o", required=True,
                        help="Directorio donde se guardarán los archivos de salida.")
    parser.add_argument("--sku-column", type=str, default=None,
                        help="Nombre de la columna que contiene los SKU (opcional).")
    parser.add_argument("--max-images", type=int, default=DEFAULT_MAX_IMAGES,
                        help=f"Máximo número de imágenes a descargar por SKU (1-10). Por defecto: {DEFAULT_MAX_IMAGES}.")
    parser.add_argument("--delay-seconds", type=float, default=DEFAULT_DELAY_SECONDS,
                        help=f"Pausa mínima entre consultas (segundos). Por defecto: {DEFAULT_DELAY_SECONDS}.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Limitar procesamiento a las primeras n filas (para pruebas).")
    parser.add_argument("--dry-run", action="store_true",
                        help="Solo buscar y extraer metadatos/URLs; no descargar imágenes ni PDFs.")
    parser.add_argument("--resume", action="store_true",
                        help="Retomar ejecución previa. Saltará SKUs ya procesados.")
    parser.add_argument("--source-map", type=str, default=None,
                        help="Archivo CSV con mapeos sku->source_url (prioridad sobre navegación).")
    parser.add_argument("--verbose", "-v", action="store_true",
                        help="Mostrar mensajes de depuración.")
    return parser


def main() -> None:
    parser = build_parser()
    args = parser.parse_args()

    if not (1 <= args.max_images <= 10):
        parser.error(f"--max-images debe estar entre 1 y 10. Valor recibido: {args.max_images}")
    if args.delay_seconds < 1.0:
        parser.error(f"--delay-seconds debe ser >= 1.0.Valor recibido: {args.delay_seconds}")

    logger = setup_logging(verbose=args.verbose)
    enricher = HikvisionEnricher(args, logger)
    exit_code = enricher.run()
    sys.exit(exit_code)


if __name__ == "__main__":
    main()
